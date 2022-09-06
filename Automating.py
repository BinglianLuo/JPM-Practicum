# -*- coding: utf-8 -*-
"""
Created on Thu Apr 21 16:29:02 2022

@author: lenovo
"""
import pandas as pd 
from scipy import optimize
from scipy.stats import norm
from scipy.special import erf,erfc
from scipy.optimize import differential_evolution
import numpy as np
import matplotlib.pyplot as plt
import pandas_market_calendars as mcal
import time
import sympy as sym
import sklearn.metrics
import seaborn as sns
from scipy.optimize import fmin_bfgs

from scipy.optimize import minimize, fsolve, root
#%%First Maturity Funcs
# BSM Option Price Calculation
def bs_price(cp_flag, s, k, t, r, v, q=0.0):
    n = norm.pd
    N = norm.cdf
    d1 = (np.log(s / k) + (r + v * v / 2.) * t) / (v * np.sqrt(t))
    d2 = d1 - v * np.sqrt(t)
    if cp_flag == 'c':
        price = s * np.exp(-q * t) * N(d1) - k * np.exp(-r * t) * N(d2)
    else:
        price = k * np.exp(-r * t) * N(-d2) - s * np.exp(-q * t) * N(-d1)
    return price


# Implied Vol Calculation
def iv_cal(cp_flag, s, k, t, r, target):
    # Need to add put formula
    def fit(sigma):
        return bs_price(cp_flag, s, k, t, r, sigma, q=0.0) - target

    if target >= s - k * np.exp(-r * t):
        # initial = np.sqrt(2 * np.pi / t) * target / s
        root = optimize.brentq(fit, -1, 6)
        return root
    else:
        print("No solution")
        return 0


# B* func Calculation
def Bfunc(S, K, t, sigma):
    inside = (K - S) / (np.sqrt(2 * t) * sigma)
    outsideup = sigma * np.sqrt(t) * np.exp(-(K - S) ** 2 / (2 * t * sigma ** 2))
    outsidedown = np.sqrt(2 * np.pi)
    return 0.5 * (S - K) * (1 - erf(inside)) + outsideup / outsidedown



# Part 3.1: Optimization over sigma0
def loss_func(s, t, k, c):
    # S: constant spot price
    # T: constant and calculated
    # K: list, imported
    # c: list, imported
    def find_sigma(x):
        a = 0
        for i in range(0, len(k)):
            a += (Bfunc(s, k[i], t, x) - c[i]) ** 2
        return a

    return optimize.fmin(find_sigma, 10, disp=False)


# Market Days Calculation
def mkt_time(t1, t2):
    nyse = mcal.get_calendar('NYSE')
    early = nyse.schedule(start_date=t1, end_date=t2)
    t = mcal.date_range(early, frequency='1D')
    return (len(t)-1) / 252

#####################################################################
# Iu Calculation
def cal_Iu(V, h):
    A = (K[0] - S0) / sigma
    B = (K[1] - S0) / sigma
    sum = 0.5 * np.exp((-h * sigma) ** 2 / 2) * \
          (erf((B + h * sigma) / np.sqrt(2)) - erf((A + h * sigma) / np.sqrt(2)))
    for k in range(0, len(dt['Strike'])):
        A = (K[k + 1] - S0) / sigma
        B = (K[k + 2] - S0) / sigma
        alpha = -V[0:k + 1].sum() - h
        beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0

        sum += 0.5 * np.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0) * \
               (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2)))
        # print(alpha, beta, beta + (alpha * sigma) ** 2 / 2 + alpha * S0,
        #       np.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0),
        #       (B - alpha * sigma) / np.sqrt(2), (A - alpha * sigma) / np.sqrt(2), sum)
    return sum

def cal_Iu1(V, h):
    A = (K[0] - S0) / sigma
    B = (K[1] - S0) / sigma
    sum = 0.5 * np.exp((-h * sigma) ** 2 / 2) * \
          (erf((B + h * sigma) / np.sqrt(2)) - erf((A + h * sigma) / np.sqrt(2)))
    for k in range(0, len(dt['Strike'])):
        A = (K[k + 1] - S0) / sigma
        B = (K[k + 2] - S0) / sigma
        alpha = -V[0:k + 1].sum() - h
        beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
        sum += 0.5 * sym.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0) * \
               (sym.erf((B - alpha * sigma) / np.sqrt(2)) - sym.erf((A - alpha * sigma) / np.sqrt(2)))
        # print(alpha, beta, beta + (alpha * sigma) ** 2 / 2 + alpha * S0,
        #       np.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0),
        #       (B - alpha * sigma) / np.sqrt(2), (A - alpha * sigma) / np.sqrt(2), sum)
    return sym.N(sum)


# Ih Calculation
# Solve for h
# Without using the value of u
def solve_Ih(V):
    def Ih(h):
        A = (K[0] - S0) / sigma
        B = (K[1] - S0) / sigma
        alpha = - h
        beta = h * S0
        sum = np.exp(beta) * \
              (2 * sigma * np.exp(alpha * S0)
               * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
               np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
               (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        for k in range(0, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += np.exp(beta) * \
                   (2 * sigma * np.exp(alpha * S0)
                    * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
                    np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
                    (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        return sum

    root = optimize.brentq(Ih, -10, 10)
    return root


def solve_Ih1(V):
    def Ih(h):
        A = (K[0] - S0) / sigma
        B = (K[1] - S0) / sigma
        alpha = - h
        beta = h * S0
        sum = np.exp(beta) * \
              (2 * sigma * np.exp(alpha * S0)
               * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
               np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
               (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        for k in range(0, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += np.exp(beta) * \
                   (2 * sigma * np.exp(alpha * S0)
                    * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
                    np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
                    (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        return sum

    root = optimize.fsolve(Ih, 0)
    return root

# Ik Calculation
def Ik(h, V, K):
    V_k = np.zeros(len(V))
    for j in range(0, len(V)):
        sum = 0
        for k in range(j, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += (2 * sigma * np.exp(alpha * S0 + beta) \
                    * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) + \
                    np.sqrt(2 * np.pi) * (alpha * sigma ** 2 - K[j + 1] + S0) * np.exp(
                        (alpha * sigma) ** 2 / 2 + alpha * S0 + beta) * \
                    (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        V_k[j] = sum / (2 * np.sqrt(2 * np.pi))
    return V_k


def Ik1(h, V, K):
    V_k = sym.zeros(1, len(V))
    for j in range(0, len(V)):
        sum = 0
        for k in range(j, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += sym.N((2 * sigma * sym.exp(alpha * S0 + beta) \
                    * (sym.exp(A * alpha * sigma - A ** 2 / 2) - sym.exp(B * alpha * sigma - B ** 2 / 2)) + \
                    sym.sqrt(2 * sym.pi) * (alpha * sigma ** 2 - K[j + 1] + S0) * sym.exp(
                        (alpha * sigma) ** 2 / 2 + alpha * S0 + beta) * \
                    (sym.erf((B - alpha * sigma) / np.sqrt(2)) - sym.erf((A - alpha * sigma) / np.sqrt(2)))))
        V_k[j] = sum / (2 * np.sqrt(2 * np.pi))
    return V_k


# f1 partial calculation
# Used for g1 partial
def f1_partial(V):
    partial = np.zeros(len(V))
    for i in range(0, len(V)):
        if delta_bid[i] <= V[i] * w[i] <= delta_ask[i]:
            partial[i] = V[i] * w[i]
        elif V[i] * w[i] > delta_ask[i]:
            partial[i] = delta_ask[i]
        else:
            partial[i] = delta_bid[i]
    return partial


# g1 partial calculation
# Not used in DE
def g1_partial1(V, u, h):
    return sym.Array(f1_partial(V) + c_mid - sym.Array(Ik1(h, V, K) * np.exp(-u))[0])


def g1_partial(V, u, h):
    return f1_partial(V) + c_mid - Ik(h, V, K) * np.exp(-u)


def sp_norm(a):
    s = 0
    for i in range (0, len(a)):
        s += a[i] ** 2
    return sym.sqrt(s)

# Sum of f1
# Used to calculate g1
def sum_f1(V):
    f = [0 for x in range(0, len(c_bid))]
    for i in range(0, len(f)):
        if (V[i] * w[i] >= delta_bid[i]) and (V[i] * w[i] <= delta_ask[i]):
            f[i] = V[i] ** 2 * w[i] / 2
        elif V[i] * w[i] > delta_ask[i]:
            f[i] = delta_ask[i] * V[i] - delta_ask[i] ** 2 / (2 * w[i])
        else:
            f[i] = delta_bid[i] * V[i] - delta_bid[i] ** 2 / (2 * w[i])
    return np.sum(f)


# g1 Calculation
def g1(u, h, V):
    res = u + sum_f1(V) + np.sum(V * c_mid) + cal_Iu(V, h) * np.exp(float(-u))
    return res

def g11(u, h, V):
    res = u + sum_f1(V) + np.sum(V * c_mid) + cal_Iu1(V, h) * np.exp(float(-u))
    return res

# DE to update V
def V_update(u, h, V, bound):
    l = len(V)
    def solve_V(v):
        return g1(u, h, v)
    bound_u = bound
    bound_l = -bound
    bounds = [(bound_l, bound_u)] * l
    result = differential_evolution(solve_V, bounds)
    v = result.x
    return v


# print(fmin_bfgs(g1, np.ones(len(dt['Strike'])) * 0.000001, fprime=g1_prime))


# Call price calculation
def call_price(u, h, V):
    return Ik(h, V, K)*np.exp(-u)


def precision (c_mid, c_model):
    return sklearn.metrics.r2_score(c_mid, c_model)


print("Complete")
print("Complete")

#%%First Maturity Outputs Func
def ivplot(u,h,V):
    #meth='mim G1'
    dt['Model'] = call_price(u, h, V)
    for i in range(0, len(dt)):
        dt.loc[i, 'Moneyness'] = dt.loc[i, 'Normalized'] / S0
        dt.loc[i, 'Bid_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Bid'])
        dt.loc[i, 'Ask_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Ask'])
        dt.loc[i, 'Mid_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Mid'])
        dt.loc[i, 'Model_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Model'])

    dt1 = dt[~dt['Bid_IV'].isin([0])]
    # dt = dt[~dt['Model_IV'].isin([0])]
    plt.scatter(dt1['Moneyness'], dt1['Ask_IV'], c='blue', marker='o', s=10)
    plt.scatter(dt1['Moneyness'], dt1['Bid_IV'], c='orange', marker='^', s=10)
    plt.plot(dt1['Moneyness'], dt1['Mid_IV'], c='purple')
    plt.scatter(dt1['Moneyness'], dt1['Model_IV'], c='red', marker='x', s=10)
    # plt.xlim(0.85, 1.2)
    #plt.ylim(0.5, 0.7)
    plt.legend(labels=['Ask', 'Bid', 'Mid', 'Model'])
    plt.title("Maturity %s, Deviation %s "%(date,np.round(deviation_val, 4)))
    plt.show()
    
def deviation_ratio(u,h,V):
    sum_value=np.sum(np.abs((call_price(u, h, V)-dt['Mid']))/(dt['Ask']-dt['Bid'])/len(dt))
    return sum_value

def min_g1(V):
    return g1(u, h, V)

print("Complete") 


#%% First Maturity Data
# Remove bid = ask
date_list = list(pd.read_excel(r'/Users/binglianluo/Desktop/Spring2022/Practicum/date_list.xlsx', header=None)[0])
date = date_list[0]
dt = pd.read_excel(r'/Users/binglianluo/Desktop/Spring2022/Practicum/Model4.xlsx', sheet_name=date)
dt['Check'] = (dt['Bid'] == dt['Ask'])
dt = dt[~dt['Check'].isin([True])]
dt = dt.reset_index(drop=True)

S0 = 1
t = mkt_time('2022-01-25', date)
r = 0.02
F = S0 * np.exp(r * t)

dt["Normalized"] = np.log(dt["Strike"] / S0) / (np.sqrt(t))
dt = dt[dt["Normalized"] >= -1.5]
dt = dt[dt["Normalized"] <= 2]
dt = dt.reset_index(drop=True)

K = np.append(0, dt['Strike'])
K = np.append(K, 10000)


dt['Bid_IV'] = 0
dt['Ask_IV'] = 0
dt['Mid'] = dt['Model']
sigma = (loss_func(S0, t, dt['Strike'], dt['Mid']) * np.sqrt(t))[0]

# Scaling
dt = dt / S0
sigma = sigma / S0
K = K / S0
dt['Normalized'] = dt['Normalized'] * S0
S0 = 1

c_bid = dt['Bid']
c_ask = dt['Ask']
c_mid = dt['Model']
delta_bid = [(a - b) for a, b in zip(c_bid, c_mid)]
delta_ask = [(a - b) for a, b in zip(c_ask, c_mid)]
# w = np.zeros(len(dt['Strike']))
w = [0.1 * (a - b) for a, b in zip(c_ask, c_bid)]

print("Complete")
#%% First Maturity Iteration
u=0;h=0;
V = np.zeros(len(dt['Strike']))

#u_l = [];h_l = [];g1_l = [];r_l = [];D_l = []
total_start_time = time.time()

for i in range(0, 20):
    start_time = time.time()
    u = float(sym.log(cal_Iu(V, h)))
    #u_l.append(u)
    h = solve_Ih(V)
    #h_l.append(h)

    # V = fsolve(min_g1partial, V)
    res = minimize(min_g1, V, method='BFGS')
    V = res.x

    #func_val = g11(u, h, V)
    #g1_l.append(func_val)
    
    deviation_val = deviation_ratio(u,h,V)
    #D_l.append(deviation_val)

    #r = np.sqrt(precision(c_mid, call_price(u, h, V)))
    #r_l.append(r)

    print("Iteration %s --- %s seconds ---" % (i + 1, time.time() - start_time))
    #print("G1 = ", func_val)
    print("u = %s , h = %s" % (u, h))
    print("Deviation Ratio = ", deviation_val)

    if deviation_val<0.08:
        break
#u_star=u;h_star=h;V_star=V
print("ToTal Iteration %s --- %s seconds ---" % (i + 1, time.time() - total_start_time))

#%% First Maturity Save
#Getting Output
#ivplot(u,h,V)
#dt['Model'] = call_price(u, h, V)

from openpyxl import load_workbook
def write_to_excel(date):
    path = r'/Users/binglianluo/Desktop/Spring2022/Practicum/Model4_1.xlsx'
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    d = dt[['Strike', 'Bid', 'Ask', 'Mid', 'Model', 'Normalized']]
    d.to_excel(writer, sheet_name=date, index=None)
    writer.save()
    return 0
#write_to_excel(date)



#%%Second Maturity 
maturity_num=14

#Moving into 1-2
def get_weight(s_l,mean,vol):
    denom = np.sum(norm.pd(s_l, mean, vol))
    numer = norm.pd(s_l, mean, vol)
    return (numer/denom )

def get_sigmat1(maturity_num):
    #maturity_num >= 1
    date = date_list[maturity_num-1]
    dt = pd.read_excel(r'/Users/binglianluo/Desktop/Spring2022/Practicum/Model4.xlsx', sheet_name=date)
    dt['Check'] = (dt['Bid'] == dt['Ask'])
    dt = dt[~dt['Check'].isin([True])]
    dt = dt.reset_index(drop=True)
    S0 = 1
    t = mkt_time('2022-01-25', date)
    dt["Normalized"] = np.log(dt["Strike"] / S0) / (np.sqrt(t))
    dt = dt[dt["Normalized"] >= -1.5]
    dt = dt[dt["Normalized"] <= 2]
    dt = dt.reset_index(drop=True)
    dt['Mid'] = dt['Model']
    sigmat1=(loss_func(S0, t, dt['Strike'], dt['Mid']) * np.sqrt(t))[0]
    sigmat1 /= S0
    return(sigmat1)


date1 = date_list[maturity_num-1]
date2 = date_list[maturity_num]


dt1 = pd.read_excel(r'/Users/binglianluo/Desktop/Spring2022/Practicum/Model4.xlsx',sheet_name=date1)
dt2 = pd.read_excel(r'/Users/binglianluo/Desktop/Spring2022/Practicum/Model4.xlsx',sheet_name=date2)

dt1['Check'] = (dt1['Bid'] == dt1['Ask'])
dt1 = dt1[~dt1['Check'].isin([True])]
dt1 = dt1.reset_index(drop=True)
dt2['Check'] = (dt2['Bid'] == dt2['Ask'])
dt2 = dt2[~dt2['Check'].isin([True])]
dt1 = dt2.reset_index(drop=True)

S0 = 1
t1 = mkt_time('2022-01-25', date1)
t2 = mkt_time('2022-01-25', date2)
r = 0.02
F = S0 * np.exp(r * t)

dt1["Normalized"] = np.log(dt1["Strike"] / S0) / (np.sqrt(t1))
dt1 = dt1[dt1["Normalized"] >= -1.5]
dt1 = dt1[dt1["Normalized"] <= 2]
dt1 = dt1.reset_index(drop=True)
dt2["Normalized"] = np.log(dt2["Strike"] / S0) / (np.sqrt(t2))
dt2 = dt2[dt2["Normalized"] >= -1.5]
dt2 = dt2[dt2["Normalized"] <= 2]
dt2 = dt2.reset_index(drop=True)

K1 = np.append(0, dt1['Strike'])
K1 = np.append(K1, 10000)
K2 = np.append(0, dt2['Strike'])
K2 = np.append(K2, 10000)
dt1['OriMid'] = dt1['Mid']
dt2['OriMid'] = dt2['Mid']
dt1['Mid'] = dt1['Model']
dt2['Mid'] = dt2['Model']

dt1['Bid_IV'] = 0
dt1['Ask_IV'] = 0
dt1['Moneyness'] = 0

# Scaling S1
S0 = 1
dt1 = dt1 / S0
K1 = K1 / S0
dt1['Normalized'] = dt1['Normalized'] * S0
S0 = 1

# Scaling S2
S0 = 1
dt2 = dt2 / S0
K2 = K2 / S0
dt2['Normalized'] = dt2['Normalized'] * S0
S0 = 1

#Discretizing S1
sigmat1=get_sigmat1(maturity_num)
#S1_list=np.linspace(S0-3*sigmat1,S0+3*sigmat1,50)
#S1_list=dt1['Strike']
def error_func(sig, kmin, n):
    # S: constant spot price
    # T: constant and calculated
    # K: list, imported
    # c: list, imported
    
    def E_func(a, sigma, Kmin, n):
        b = 2 * S0 - a
        A1 = 1 / (2 * sigma ** 2)
        m = S0 - Kmin
        err1=(b-a)/n
        err2=-1/(2*np.sqrt(2*np.pi)*sigma*A1)*np.exp(-A1*(b-S0)**2)
        err3=m/(2*np.sqrt(2*np.pi)*sigma*A1)*erfc(np.sqrt(A1)*(b-S0))
        return (err1+err2+err3)
    a = kmin
    def find_a(x):
        e = E_func(x , sig , kmin, n)
        return e
    ini = kmin
    return optimize.minimize(find_a, ini)
def get_discrete(a,n):
    b=2*S0-a
    s_l=np.linspace(a,b,n)
    return s_l
#a=200/918.4
#n=50
#result = error_func(sigmat1 , dt1['Strike'][0], n)
#a = result.x[0]
#S1_list=get_discrete(a,n)

S1_list=np.linspace(list(dt1['Strike'])[0],2*S0-list(dt1['Strike'])[0],50)
m0_list=norm.pd(S1_list, S0, sigmat1)
weight_list = get_weight(S1_list, S0, sigmat1)

# Part 3.1: Optimization over sigma0
def loss_func_list(s_l, w_l, t, k, c):
    # S: list, price
    # T: constant and calculated
    # K: list, imported
    # c: list, imported
    def find_sigma(x):
        # x: [sigma0 , beta]
        #sig: list
        sig = x[0] * s_l ** x[1]
        a = 0
        for i in range(len(k)):
            expectation = 0
            for j in range(len(s_l)):
                expectation += Bfunc(s_l[j], k[i], t, sig[j])*w_l[j]
            a += (expectation - c[i]) ** 2
        return a
    
    ini = [0.5,0.5]
    return optimize.minimize(find_sigma, ini)

X=loss_func_list(S1_list,weight_list,(t2-t1),dt2['Strike'],dt2['Mid'])
sigmat1_list = X.x[0] * S1_list ** X.x[1] * np.sqrt(t2-t1)
#X.x[0] * 1 ** X.x[1] * np.sqrt(t2-t1) 


#%% Second Maturity Funcs
# Iu Calculation
def new_cal_Iu(V, h):
    A = (K[0] - S0) / sigma
    B = (K[1] - S0) / sigma
    sum = 0.5 * np.exp((-h * sigma) ** 2 / 2) * \
          (erf((B + h * sigma) / np.sqrt(2)) - erf((A + h * sigma) / np.sqrt(2)))
    for k in range(0, len(dt['Strike'])):
        A = (K[k + 1] - S0) / sigma
        B = (K[k + 2] - S0) / sigma
        alpha = -np.sum(V[0:k + 1]) - h
        beta = np.sum((V[0:k + 1] * K[1:k + 2])) + h * S0

        sum += 0.5 * np.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0) * \
               (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2)))
        # print(alpha, beta, beta + (alpha * sigma) ** 2 / 2 + alpha * S0,
        #       np.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0),
        #       (B - alpha * sigma) / np.sqrt(2), (A - alpha * sigma) / np.sqrt(2), sum)
    return sum
def expec_cal_Iu(V,h):
    Iu_sum=0
    for i in range(len(S1_list)):
        Iu_sum += new_cal_Iu(V, h, sigmat1_list[i], K2, dt2, S1_list[i])*weight_list[i]
    return Iu_sum

def new_cal_Iu1(V, h):
    A = (K[0] - S0) / sigma
    B = (K[1] - S0) / sigma
    sum = 0.5 * np.exp((-h * sigma) ** 2 / 2) * \
          (erf((B + h * sigma) / np.sqrt(2)) - erf((A + h * sigma) / np.sqrt(2)))
    for k in range(0, len(dt['Strike'])):
        A = (K[k + 1] - S0) / sigma
        B = (K[k + 2] - S0) / sigma
        alpha = -V[0:k + 1].sum() - h
        beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
        sum += 0.5 * sym.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0) * \
               (sym.erf((B - alpha * sigma) / np.sqrt(2)) - sym.erf((A - alpha * sigma) / np.sqrt(2)))
        # print(alpha, beta, beta + (alpha * sigma) ** 2 / 2 + alpha * S0,
        #       np.exp(beta + (alpha * sigma) ** 2 / 2 + alpha * S0),
        #       (B - alpha * sigma) / np.sqrt(2), (A - alpha * sigma) / np.sqrt(2), sum)
    return sym.N(sum)
def expec_cal_Iu1(V,h):
    Iu_sum=0
    for i in range(len(S1_list)):
        Iu_sum += new_cal_Iu1(V, h, sigmat1_list[i], K2, dt2, S1_list[i])*weight_list[i]
    return Iu_sum


# Ih Calculation
# Solve for h
# Without using the value of u
def new_solve_Ih(V):
    def Ih(h):
        A = (K[0] - S0) / sigma
        B = (K[1] - S0) / sigma
        alpha = - h
        beta = h * S0
        sum = np.exp(beta) * \
              (2 * sigma * np.exp(alpha * S0)
               * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
               np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
               (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        for k in range(0, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += np.exp(beta) * \
                   (2 * sigma * np.exp(alpha * S0)
                    * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
                    np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
                    (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        return sum        
    root = optimize.brentq(Ih, -10, 10)
    return root


def new_solve_Ih1(V):
    def Ih(h):
        A = (K[0] - S0) / sigma
        B = (K[1] - S0) / sigma
        alpha = - h
        beta = h * S0
        sum = np.exp(beta) * \
              (2 * sigma * np.exp(alpha * S0)
               * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
               np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
               (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        for k in range(0, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += np.exp(beta) * \
                   (2 * sigma * np.exp(alpha * S0)
                    * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) +
                    np.sqrt(2 * np.pi) * alpha * sigma ** 2 * np.exp((alpha * sigma) ** 2 / 2 + alpha * S0) *
                    (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        return sum
    root = optimize.fsolve(Ih, 0)
    return root



# Ik Calculation
def new_Ik(h, V, K):
    V_k = np.zeros(len(V))
    for j in range(0, len(V)):
        sum = 0
        for k in range(j, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += (2 * sigma * np.exp(alpha * S0 + beta) \
                    * (np.exp(A * alpha * sigma - A ** 2 / 2) - np.exp(B * alpha * sigma - B ** 2 / 2)) + \
                    np.sqrt(2 * np.pi) * (alpha * sigma ** 2 - K[j + 1] + S0) * np.exp(
                        (alpha * sigma) ** 2 / 2 + alpha * S0 + beta) * \
                    (erf((B - alpha * sigma) / np.sqrt(2)) - erf((A - alpha * sigma) / np.sqrt(2))))
        V_k[j] = sum / (2 * np.sqrt(2 * np.pi))
    return V_k


def new_Ik1(h, V, K):
    V_k = sym.zeros(1, len(V))
    for j in range(0, len(V)):
        sum = 0
        for k in range(j, len(dt['Strike'])):
            A = (K[k + 1] - S0) / sigma
            B = (K[k + 2] - S0) / sigma
            alpha = -V[0:k + 1].sum() - h
            beta = (V[0:k + 1] * K[1:k + 2]).sum() + h * S0
            sum += sym.N((2 * sigma * sym.exp(alpha * S0 + beta) \
                    * (sym.exp(A * alpha * sigma - A ** 2 / 2) - sym.exp(B * alpha * sigma - B ** 2 / 2)) + \
                    sym.sqrt(2 * sym.pi) * (alpha * sigma ** 2 - K[j + 1] + S0) * sym.exp(
                        (alpha * sigma) ** 2 / 2 + alpha * S0 + beta) * \
                    (sym.erf((B - alpha * sigma) / np.sqrt(2)) - sym.erf((A - alpha * sigma) / np.sqrt(2)))))
        V_k[j] = sum / (2 * np.sqrt(2 * np.pi))
    return V_k


###
c_bid = dt2['Bid']
c_ask = dt2['Ask']
c_mid = dt2['Model']
delta_bid = [(a - b) for a, b in zip(c_bid, c_mid)]
delta_ask = [(a - b) for a, b in zip(c_ask, c_mid)]
w = [0.01 * (a - b) for a, b in zip(c_ask, c_bid)]
# Sum of f1
# Used to calculate g1
def sum_f1(V):
    f = [0 for x in range(0, len(c_bid))]
    for i in range(0, len(f)):
        if (V[i] * w[i] >= delta_bid[i]) and (V[i] * w[i] <= delta_ask[i]):
            f[i] = V[i] ** 2 * w[i] / 2
        elif V[i] * w[i] > delta_ask[i]:
            f[i] = delta_ask[i] * V[i] - delta_ask[i] ** 2 / (2 * w[i])
        else:
            f[i] = delta_bid[i] * V[i] - delta_bid[i] ** 2 / (2 * w[i])
    return np.sum(f)

# g1 Calculation
def g1(u, h, V):
    res = u + sum_f1(V) + np.sum(V * c_mid) + new_cal_Iu(V, h) * np.exp(float(-u))
    return res
def g11(u, h, V):
    res = u + sum_f1(V) + np.sum(V * c_mid) + new_cal_Iu1(V, h) * np.exp(float(-u))
    return res


def call_price(u, h, V):
    return new_Ik(h, V, K)*np.exp(-u)

print('Complete')
print('Complete')

#%% Second Maturity Output

def ivplot(u,h,V):
    #meth='mim G1'
    dt['Model'] = new_Ik(h, V, K)*np.exp(-u)
    S0=1
    #dt['Model'] = new_call_price(Iu_list,Ik_list,weight_list)
    for i in range(0, len(dt)):
        dt.loc[i, 'Moneyness'] = dt.loc[i, 'Normalized'] / S0
        dt.loc[i, 'Bid_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Bid'])
        dt.loc[i, 'Ask_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Ask'])
        dt.loc[i, 'Mid_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Mid'])
        dt.loc[i, 'Model_IV'] = iv_cal('c', S0, dt.loc[i, 'Strike'], t, 0, dt.loc[i, 'Model'])


    dt1 = dt[~dt['Bid_IV'].isin([0])]
    # dt = dt[~dt['Model_IV'].isin([0])]
    plt.scatter(dt1['Moneyness'], dt1['Ask_IV'], c='blue', marker='o', s=10)
    plt.scatter(dt1['Moneyness'], dt1['Bid_IV'], c='orange', marker='^', s=10)
    plt.plot(dt1['Moneyness'], dt1['Mid_IV'], c='purple')
    plt.scatter(dt1['Moneyness'], dt1['Model_IV'], c='red', marker='x', s=10)
    # plt.xlim(0.85, 1.2)
    #plt.ylim(2, 4)
    plt.legend(labels=['Mid', 'Ask', 'Bid', 'Model'])
    #plt.title("TSLA, Mat = 3-6 Days, Meth=P*, S1( %s )" %(n+1))
    plt.title('Maturity %s, Deviation %s '%(date2,np.round(deviation_val, 4))) 
    plt.show()

def min_g1(V):
    return g1(u, h, V)

def deviation_ratio(u,h,V):
    sum_value=np.sum(np.abs((call_price(u, h, V)-dt['OriMid']))/(dt['Ask']-dt['Bid'])/len(dt))
    return sum_value

print('Complete')

#%% Second Maturity Iteration
K=K2;dt=dt2;t=t2-t1;
total_start_time = time.time()
V=np.zeros(len(dt2))
u_l=np.zeros(len(S1_list));h_l=np.zeros(len(S1_list))
u=0;h=0;

for i in range(100):
    print(i)
    start_time = time.time()
    for s in range(len(S1_list)):
        #print(s)
        sigma=sigmat1_list[s];S0=S1_list[s]
        u_l[s] = float(sym.log(new_cal_Iu1(V,h_l[s])))
        h_l[s] = new_solve_Ih1(V)[0]
        # u_l[s] = np.log(new_cal_Iu(V, h_l[s]))#/norm.cdf(s,1,sigmat1)
        # h_l[s] = new_solve_Ih(V)
        
    S0=1
    u=np.average(u_l);h=np.average(h_l)
    
    res = minimize(min_g1, V, method='L-BFGS-B')
    V = res.x
    #func_val = g11(u, h, V)
    deviation_val = deviation_ratio(u,h,V)

    print("Iteration %s --- %s seconds ---" % (i + 1, time.time() - start_time))
    #print("G1 = ", func_val)
    print("u = %s , h = %s" % (u, h))
    print("Deviation Ratio = ", deviation_val)

    if deviation_val<0.14:
        break
#u_star=u;h_star=h;V_star=V
print("ToTal Iteration %s --- %s seconds ---" % (i + 1, time.time() - total_start_time))

#%% Second Maturity Save
from openpyxl import load_workbook
ivplot(u,h,V)
dt['Model'] = call_price(u, h, V)
write_to_excel(date2)  
